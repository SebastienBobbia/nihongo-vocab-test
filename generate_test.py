# =============================================================================
# generate_test.py — Génération automatique de tests de vocabulaire japonais
# =============================================================================
# Usage:
#   python generate_test.py                         → génère tous les tests
#   python generate_test.py --profile bob_N4        → profil spécifique
#   python generate_test.py --profile bob_N4 --sheets N4-14 N4-15
#
# Format du test :
#   - Feuille 1 "Kanji->Hiragana" : Kanji affiché, choisir la bonne lecture
#   - Feuille 2 "FR->JP"         : Mot français affiché, choisir le bon kanji
#   - 50 questions par feuille, 5 questions × 10 colonnes
# =============================================================================

import argparse
import os
import random
import sys

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import PROFILES, QUESTIONS_PER_TEST


# =============================================================================
# Chargement des données
# =============================================================================

def detect_sheets(vocab_file: str, level: str) -> list[str]:
    """Retourne toutes les feuilles correspondant au niveau (ex: N4-1, N4-2, ...)."""
    wb = openpyxl.load_workbook(vocab_file, read_only=True, data_only=True)
    sheets = sorted(
        [s for s in wb.sheetnames if s.startswith(f"{level}-")],
        key=lambda s: int(s.split("-")[1])
    )
    wb.close()
    return sheets


def load_vocabulary(vocab_file: str, sheets: list[str]) -> pd.DataFrame:
    """
    Charge et concatène le vocabulaire de plusieurs feuilles.
    Colonnes attendues (indices 4, 5, 6): Kanji, Hiragana, Francais.
    Les lignes avec des valeurs manquantes sont supprimées.
    """
    frames = []
    for sheet in sheets:
        df = pd.read_excel(
            vocab_file,
            sheet_name=sheet,
            usecols=[4, 5, 6],
            header=None,
            names=["Kanji", "Hiragana", "Francais"],
            skiprows=1,
        )
        df["_sheet"] = sheet
        frames.append(df)

    df_all = pd.concat(frames, ignore_index=True)
    df_all = df_all.dropna(subset=["Kanji", "Hiragana", "Francais"])
    df_all = df_all.reset_index(drop=True)
    return df_all


# =============================================================================
# Sélection des mauvaises réponses
# =============================================================================

def get_wrong_choices(correct_value: str, column: str, df: pd.DataFrame, n: int = 3) -> list:
    """
    Sélectionne n mauvaises réponses dans le voisinage (groupe de 10) de la bonne réponse.
    - Évite les doublons avec la bonne réponse.
    - Si le groupe est trop petit, élargit à tout le DataFrame.
    """
    matches = df[df[column] == correct_value]
    if matches.empty:
        # Fallback: piocher n'importe où
        pool = df[df[column] != correct_value][column].unique().tolist()
        return random.sample(pool, min(n, len(pool)))

    index_correct = matches.index[0]
    group_start = (index_correct // 10) * 10
    group_end = group_start + 9

    group = df.iloc[group_start : group_end + 1]
    pool = list(set(group[column].tolist()) - {correct_value})

    # Si pas assez dans le groupe, compléter avec le reste du DataFrame
    if len(pool) < n:
        extra = list(set(df[column].tolist()) - {correct_value} - set(pool))
        random.shuffle(extra)
        pool += extra

    return random.sample(pool, min(n, len(pool)))


# =============================================================================
# Construction du format Excel (grille vide)
# =============================================================================

def _build_empty_grid(ws) -> None:
    """
    Crée la grille de 50 cases (10 colonnes × 5 lignes) avec bordures et numéros.
    Layout : colonnes par triplets (num | réponses | séparateur), lignes par groupes de 6.
    """
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    fill_separator = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    fill_ref = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")

    # Largeurs de colonnes
    for col in range(1, 29, 3):   # colonnes numéro  (A, D, G, ...)
        ws.column_dimensions[get_column_letter(col)].width = 3.8
    for col in range(2, 30, 3):   # colonnes réponse (B, E, H, ...)
        ws.column_dimensions[get_column_letter(col)].width = 15.8
    for col in range(3, 31, 3):   # colonnes séparateur (C, F, I, ...)
        ws.column_dimensions[get_column_letter(col)].width = 1.8

    # Hauteurs des lignes séparatrices
    for row in range(6, 31, 6):
        ws.row_dimensions[row].height = 10

    # Couleur des séparateurs
    for row in range(6, 31, 6):
        for col in range(1, 31):
            ws[f"{get_column_letter(col)}{row}"].fill = fill_separator
    for col in range(3, 31, 3):
        for row in range(1, 31):
            ws[f"{get_column_letter(col)}{row}"].fill = fill_separator

    # Référence couleur de sélection en bas
    ws[f"{get_column_letter(2)}31"].fill = fill_ref

    counter = 1
    for col in range(1, 29, 3):       # 10 triplets de colonnes
        for row in range(1, 26, 6):   # 5 blocs de 5 lignes

            letter_num  = get_column_letter(col)
            letter_ans  = get_column_letter(col + 1)

            # Bordures fines intérieures
            for r in range(row, row + 5):
                for c in range(col, col + 2):
                    ws[f"{get_column_letter(c)}{r}"].border = thin

            # Fusion de la cellule numéro
            ws.merge_cells(f"{letter_num}{row}:{letter_num}{row + 4}")

            # Bordures épaisses extérieures
            thick_l = Side(style="thick")
            thick_r = Side(style="thick")
            thick_t = Side(style="thick")
            thick_b = Side(style="thick")

            for r in range(row, row + 5):
                ws[f"{letter_num}{r}"].border = Border(left=thick_l)
                ws[f"{letter_ans}{r}"].border = Border(right=thick_r)
            for c in range(col, col + 2):
                ws[f"{get_column_letter(c)}{row}"].border    = Border(top=thick_t)
                ws[f"{get_column_letter(c)}{row + 4}"].border = Border(bottom=thick_b)

            # Coins — combiner les côtés
            ws[f"{letter_num}{row}"].border        = Border(top=thick_t,  left=thick_l)
            ws[f"{letter_num}{row + 1}"].border    = Border(left=thick_l, right=Side(style="thick"))
            ws[f"{letter_num}{row + 2}"].border    = Border(left=thick_l, right=Side(style="thick"))
            ws[f"{letter_num}{row + 3}"].border    = Border(left=thick_l, right=Side(style="thick"))
            ws[f"{letter_num}{row + 4}"].border    = Border(bottom=thick_b, left=thick_l, right=Side(style="thick"))
            ws[f"{letter_ans}{row}"].border        = Border(top=thick_t,  right=thick_r)
            ws[f"{letter_ans}{row + 4}"].border    = Border(bottom=thick_b, right=thick_r)

            # Numéro de question
            ws[f"{letter_num}{row}"].value     = counter
            ws[f"{letter_num}{row}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{letter_num}{row}"].font      = Font(size=14)

            # Encadrement épais de la cellule question (ligne du haut)
            ws[f"{letter_ans}{row}"].border = Border(
                top=Side(style="thick"), right=Side(style="thick")
            )

            counter += 1


# =============================================================================
# Remplissage des questions
# =============================================================================

def _fill_questions(ws, questions: pd.DataFrame, question_col: str, answer_col: str,
                    df_all: pd.DataFrame) -> None:
    """
    Remplit la feuille avec les questions et les 4 propositions de réponse.
    - question_col : colonne du texte affiché en gras (la question)
    - answer_col   : colonne de la bonne réponse (parmi les 4 choix)
    """
    for i, row_data in questions.iterrows():
        question_text  = row_data[question_col]
        correct_answer = row_data[answer_col]

        wrong = get_wrong_choices(correct_answer, answer_col, df_all)
        choices = [correct_answer] + wrong
        random.shuffle(choices)

        # Position dans la grille
        col_idx   = (i // 5) * 3 + 2   # colonne Excel (B=2, E=5, H=8, ...)
        row_start = (i % 5) * 6 + 1    # ligne Excel (1, 7, 13, 19, 25)

        # Question (gras)
        cell = ws.cell(row=row_start, column=col_idx, value=question_text)
        cell.font = Font(bold=True)

        # 4 propositions
        for j, choice in enumerate(choices):
            ws.cell(row=row_start + j + 1, column=col_idx, value=choice)


# =============================================================================
# Génération d'un test pour UNE feuille de vocabulaire
# =============================================================================

def generate_test_for_sheet(vocab_file: str, sheet: str, output_path: str,
                             n_questions: int = QUESTIONS_PER_TEST) -> None:
    """
    Génère un fichier Excel de test pour une feuille de vocabulaire donnée.
    - Feuille 1 : Kanji  → Hiragana
    - Feuille 2 : Français → Kanji
    """
    df = load_vocabulary(vocab_file, [sheet])

    available = len(df)
    if available < n_questions:
        print(f"  [AVERTISSEMENT] {sheet} : seulement {available} mots disponibles "
              f"(demandé: {n_questions}). Ajustement automatique.")
        n_questions = available // 2  # moitié pour chaque sens

    # ---- Sélection des questions ----
    # Les deux tests ne doivent pas utiliser les mêmes mots
    # On déduplique d'abord sur la colonne Kanji pour éviter les doublons
    df_unique = df.drop_duplicates(subset=["Kanji"]).reset_index(drop=True)
    available_unique = len(df_unique)

    n_q = min(n_questions, available_unique // 2)
    if n_q < n_questions:
        print(f"  [AVERTISSEMENT] {sheet} : {available_unique} mots uniques, "
              f"chaque test réduit à {n_q} questions.")

    sample1 = df_unique.sample(n=n_q).reset_index(drop=True)
    used_kanjis = set(sample1["Kanji"])
    remaining = df_unique[~df_unique["Kanji"].isin(used_kanjis)]
    sample2 = remaining.sample(n=min(n_q, len(remaining))).reset_index(drop=True)

    # ---- Construction du classeur ----
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Kanji->Hiragana"
    _build_empty_grid(ws1)
    _fill_questions(ws1, sample1, question_col="Kanji", answer_col="Hiragana", df_all=df)

    ws2 = wb.copy_worksheet(ws1)
    ws2.title = "FR->JP"
    _fill_questions(ws2, sample2, question_col="Francais", answer_col="Kanji", df_all=df)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"  OK  {os.path.basename(output_path)}")


# =============================================================================
# Point d'entrée
# =============================================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Génère des tests de vocabulaire japonais depuis un fichier Excel."
    )
    parser.add_argument(
        "--profile", "-p",
        choices=list(PROFILES.keys()),
        default=None,
        help="Profil à utiliser (défaut: tous les profils disponibles).",
    )
    parser.add_argument(
        "--sheets", "-s",
        nargs="+",
        default=None,
        metavar="SHEET",
        help="Feuilles spécifiques à traiter (ex: N4-14 N4-15). Par défaut: toutes.",
    )
    parser.add_argument(
        "--questions", "-q",
        type=int,
        default=QUESTIONS_PER_TEST,
        help=f"Nombre de questions par test (défaut: {QUESTIONS_PER_TEST}).",
    )
    return parser.parse_args()


def run_profile(profile_name: str, profile: dict, sheets_override: list | None,
                n_questions: int) -> None:
    vocab_file  = profile["vocab_file"]
    output_dir  = profile["output_dir"]
    level       = profile["level"]

    if not os.path.isfile(vocab_file):
        print(f"[ERREUR] Fichier source introuvable : {vocab_file}")
        return

    sheets = sheets_override or profile.get("sheets") or detect_sheets(vocab_file, level)
    if not sheets:
        print(f"[ERREUR] Aucune feuille '{level}-x' trouvée dans {vocab_file}")
        return

    print(f"\n=== Profil : {profile_name} ({len(sheets)} feuille(s)) ===")
    for sheet in sheets:
        sheet_num = sheet.split("-")[1]
        output_path = os.path.join(output_dir, f"{level}-{sheet_num}_voc_test.xlsx")
        print(f"  Génération de {sheet}...", end=" ", flush=True)
        try:
            generate_test_for_sheet(vocab_file, sheet, output_path, n_questions)
        except Exception as e:
            print(f"\n  [ERREUR] {sheet} : {e}")


def main():
    args = parse_args()

    profiles_to_run = (
        {args.profile: PROFILES[args.profile]}
        if args.profile
        else PROFILES
    )

    for profile_name, profile in profiles_to_run.items():
        run_profile(profile_name, profile, args.sheets, args.questions)

    print("\nTerminé.")


if __name__ == "__main__":
    main()
