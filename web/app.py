"""
FastAPI application for Japanese vocabulary test generator
Accessible from iOS via web interface on NAS
"""

import os
import sys
import json
import tempfile
import random
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, HTMLResponse
from pydantic import BaseModel
import pandas as pd

# Add parent directory to path to import config and modules
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import PROFILES
from generate_test import generate_test_for_sheet, detect_sheets, load_vocabulary
from correct_test import correct_test_file, load_answer_key

app = FastAPI(title="Nihongo Vocab Test")

# Serve static files
static_dir = Path(__file__).parent / "static"
static_dir.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")


# ============================================================================
# Models
# ============================================================================

class TestRequest(BaseModel):
    """Request to generate a new test"""
    profile: str  # "N4" or "N5"
    sheets: Optional[list] = None  # Optional specific sheets, e.g. ["N4-14", "N4-15"]


class TestCorrectRequest(BaseModel):
    """Request to correct a test"""
    profile: str
    test_name: str
    answers: dict  # {question_number: answer_choice}


# ============================================================================
# Endpoints
# ============================================================================

@app.get("/", response_class=HTMLResponse)
async def root():
    """Serve the main HTML interface"""
    html_path = Path(__file__).parent / "static" / "index.html"
    if html_path.exists():
        return FileResponse(html_path)
    return "<h1>Nihongo Vocab Test App</h1><p>Web interface loading...</p>"


@app.get("/api/profiles")
async def get_profiles():
    """Get available profiles (N4, N5)"""
    profiles = list(PROFILES.keys())
    return {"profiles": profiles}


@app.post("/api/generate")
async def generate(request: TestRequest):
    """Generate a new test"""
    if request.profile not in PROFILES:
        raise HTTPException(status_code=400, detail=f"Unknown profile: {request.profile}")
    
    profile_config = PROFILES[request.profile]
    
    try:
        # Get sheets to generate (use provided or auto-detect)
        sheets_to_generate = request.sheets or detect_sheets(
            profile_config["vocab_file"],
            profile_config["level"]
        )
        
        if not sheets_to_generate:
            raise ValueError("No sheets found to generate")
        
        # Generate test for each sheet and collect all questions
        all_vocab_items = []
        sheet_names = []
        
        for sheet in sheets_to_generate[:2]:  # Limit to 2 sheets for web interface
            try:
                # Load vocabulary for this sheet
                vocab_df = load_vocabulary(profile_config["vocab_file"], [sheet])
                
                # Store sheet info
                sheet_names.append(sheet)
                
                # Split for two question types (use first half for kanji->kana, second for fr->jp)
                mid = len(vocab_df) // 2
                
                all_vocab_items.append({
                    "sheet": sheet,
                    "kanji_kana_df": vocab_df.iloc[:mid],
                    "fr_jp_df": vocab_df.iloc[mid:]
                })
            except Exception as e:
                print(f"Warning: Could not process sheet {sheet}: {e}")
                continue
        
        if not all_vocab_items:
            raise ValueError("Failed to generate any questions")
        
        # Create test data for web interface
        test_data = prepare_test_for_web_from_vocab(all_vocab_items)
        
        return {
            "success": True,
            "test_data": test_data,
            "sheet_names": sheet_names,
            "total_questions": len(test_data.get("questions", []))
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/correct")
async def correct(request: TestCorrectRequest):
    """Correct a test based on provided answers"""
    if request.profile not in PROFILES:
        raise HTTPException(status_code=400, detail=f"Unknown profile: {request.profile}")
    
    try:
        # Create a temporary test file with answers marked
        # This will be passed to the correction function
        correction_data = prepare_correction(
            profile=request.profile,
            test_name=request.test_name,
            answers=request.answers
        )
        
        return {
            "success": True,
            "correction": correction_data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/available-sheets/{profile}")
async def get_available_sheets(profile: str):
    """Get list of available sheet names for a profile"""
    if profile not in PROFILES:
        raise HTTPException(status_code=400, detail=f"Unknown profile: {profile}")
    
    profile_config = PROFILES[profile]
    vocab_file = profile_config["vocab_file"]
    
    if not Path(vocab_file).exists():
        raise HTTPException(status_code=404, detail=f"Vocabulary file not found: {vocab_file}")
    
    try:
        import pandas as pd
        xls = pd.ExcelFile(vocab_file)
        sheets = [s for s in xls.sheet_names if s.startswith(profile_config["level"])]
        return {"sheets": sorted(sheets)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Helper Functions
# ============================================================================

def prepare_test_for_web_from_vocab(vocab_data: list) -> dict:
    """
    Convert vocabulary data to test questions for web interface
    """
    
    questions = []
    question_id = 1
    
    for vocab_item in vocab_data:
        sheet = vocab_item["sheet"]
        
        # Generate Kanji → Hiragana questions
        for idx, (_, row) in enumerate(vocab_item["kanji_kana_df"].iterrows()):
            if idx >= 25:  # Limit to 25 questions per type
                break
                
            kanji = str(row.get('Kanji', '')).strip()
            hiragana = str(row.get('Hiragana', '')).strip()
            
            if not kanji or not hiragana:
                continue
            
            # Create choices - for now, we'll use placeholder choices
            # In production, these should be wrong distractors from other vocab
            choices = [hiragana, f"選択肢{question_id % 4}", f"選択肢{(question_id+1) % 4}", f"選択肢{(question_id+2) % 4}"]
            random.shuffle(choices)
            correct_idx = choices.index(hiragana)
            
            questions.append({
                "id": question_id,
                "type": "kanji_kana",
                "sheet": sheet,
                "question": kanji,
                "choices": choices,
                "correct_index": correct_idx
            })
            question_id += 1
        
        # Generate French → Japanese questions
        for idx, (_, row) in enumerate(vocab_item["fr_jp_df"].iterrows()):
            if idx >= 25:  # Limit to 25 questions per type
                break
                
            french = str(row.get('Francais', '')).strip()
            kanji = str(row.get('Kanji', '')).strip()
            
            if not french or not kanji:
                continue
            
            # Create choices
            choices = [kanji, f"選択肢{question_id % 4}", f"選択肢{(question_id+1) % 4}", f"選択肢{(question_id+2) % 4}"]
            random.shuffle(choices)
            correct_idx = choices.index(kanji)
            
            questions.append({
                "id": question_id,
                "type": "fr_jp",
                "sheet": sheet,
                "question": french,
                "choices": choices,
                "correct_index": correct_idx
            })
            question_id += 1
    
    return {"questions": questions}


def prepare_test_for_web(test_file: Path, sheet_names: list, profile: str) -> dict:
    """
    Convert Excel test file to JSON for web interface
    Returns structure like:
    {
        "sections": [
            {
                "name": "Kanji→Hiragana",
                "questions": [
                    {
                        "id": 1,
                        "question": "漢字",
                        "choices": ["ひらがな1", "ひらがな2", "ひらがな3", "ひらがな4"],
                        "correct_answer": 0  # only for review
                    }
                ]
            }
        ]
    }
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.load_workbook(test_file)
        sections = []
        
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            questions = []
            
            # Read questions from sheet (assuming structure: question in col A, choices in B-E)
            row = 2  # Start from row 2 (row 1 is header)
            question_id = 1
            
            while True:
                question_cell = ws[f"A{row}"]
                if not question_cell.value:
                    break
                
                choices = []
                for col in ["B", "C", "D", "E"]:
                    cell_value = ws[f"{col}{row}"].value
                    if cell_value:
                        choices.append(str(cell_value))
                
                if len(choices) == 4:
                    questions.append({
                        "id": question_id,
                        "question": str(question_cell.value),
                        "choices": choices,
                        "user_answer": None
                    })
                    question_id += 1
                
                row += 1
            
            if questions:
                sections.append({
                    "name": sheet_name,
                    "questions": questions
                })
        
        return {"sections": sections}
    
    except Exception as e:
        # Fallback: return empty structure
        return {"sections": [], "error": str(e)}


def prepare_correction(profile: str, test_name: str, answers: dict) -> dict:
    """
    Process answers and return correction results
    """
    # This is a simplified version; in real implementation,
    # you'd compare against the Excel file's correct answers
    
    return {
        "test_name": test_name,
        "total_questions": len(answers),
        "correct_answers": 0,
        "score": 0,
        "details": []
    }


# ============================================================================
# Health check
# ============================================================================

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "ok"}
