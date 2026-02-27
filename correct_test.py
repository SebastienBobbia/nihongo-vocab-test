# =============================================================================
# correct_test.py — Correction automatique des tests de vocabulaire japonais
# =============================================================================
# Usage:
#   python correct_test.py --test N4-14 --profile bob_N4
#   python correct_test.py --profile bob_N4         → corrige tous les tests du profil
#
# Le script :
#   1. Charge le fichier test rempli par l'élève (_voc_test.xlsx)
#   2. Compare les réponses sélectionnées (surlignées en bleu) avec les bonnes réponses
#   3. Colore le numéro de question : vert (correct) / rouge (incorrect)
#   4. Encadre la bonne réponse en vert si la réponse est fausse
#   5. Crée une feuille "Correction" avec le score détaillé
#   6. Sauvegarde dans un nouveau fichier _corrige.xlsx
# =============================================================================

import argparse
import os
import sys

import openpyxl
from openpyxl.styles import Border, PatternFill, Side

from config import PROFILES, SELECTION_COLOR


# =============================================================================
# Styles
# =============================================================================

FILL_CORRECT   = PatternFill(start_color="00C851", end_color="00C851", fill_type="solid")
FILL_INCORRECT = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
BORDER_CORRECT_ANSWER = Border(
    left=Side(border_style="thick", color="00C851"),
    right=Side(border_style="thick", color="00C851"),
    top=Side(border_style="thick", color="00C851"),
    bottom=Side(border_style="thick", color="00C851"),
)

# Positions fixes de toutes les questions dans la grille (10 col × 5 lignes)
QUESTION_POSITIONS = [
    (row, col)
    for row in [1, 7, 13, 19, 25]
    for col in range(2, 32, 3)
]


# =============================================================================
# Chargement du référentiel de réponses correctes
# =============================================================================

def load_answer_key(vocab_file: str, sheets: list[str]) -> dict[str, str]:
    """
    Construit deux dictionnaires fusionnés :
      - Kanji    → Hiragana  (pour la feuille Kanji->Hiragana)
      - Français → Kanji     (pour la feuille FR->JP)
    Retourne un dict unique question_text → bonne_réponse.
    """
    wb = openpyxl.load_workbook(vocab_file, data_only=True, read_only=True)
    key: dict[str, str] = {}

    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            print(f"  [AVERTISSEMENT] Feuille '{sheet_name}' absente de {vocab_file}")
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            if len(row) < 7:
                continue
            kanji, hiragana, francais = row[4], row[5], row[6]
            if kanji and hiragana:
                key[kanji]   = hiragana   # Kanji  → Hiragana
            if francais and kanji:
                key[francais] = kanji     # Français → Kanji

    wb.close()
    return key


# =============================================================================
# Correction d'une feuille de test
# =============================================================================

def _correct_sheet(ws, answer_key: dict[str, str]) -> tuple[int, int]:
    """
    Corrige toutes les questions d'une feuille.
    Retourne (score, total_questions_avec_réponse_détectée).
    """
    score = 0
    total = 0

    for row, col in QUESTION_POSITIONS:
        question_cell = ws.cell(row=row, column=col)
        number_cell   = ws.cell(row=row, column=col - 1)
        question_text = question_cell.value

        if not question_text or question_text not in answer_key:
            continue  # question vide ou inconnue

        total += 1
        correct_answer = answer_key[question_text]

        # Trouver la cellule contenant la bonne réponse
        correct_cell = None
        for i in range(1, 5):
            cell = ws.cell(row=row + i, column=col)
            if cell.value == correct_answer:
                correct_cell = cell
                break

        if correct_cell is None:
            continue  # bonne réponse absente du test (ne devrait pas arriver)

        # Trouver la réponse sélectionnée par l'élève (cellule bleue)
        selected_cell = None
        for i in range(1, 5):
            cell = ws.cell(row=row + i, column=col)
            try:
                color = cell.fill.start_color.index
            except Exception:
                color = None
            if color == SELECTION_COLOR:
                selected_cell = cell
                break

        if selected_cell is None:
            continue  # aucune réponse sélectionnée → ne pas pénaliser

        if selected_cell.coordinate == correct_cell.coordinate:
            number_cell.fill = FILL_CORRECT
            score += 1
        else:
            number_cell.fill  = FILL_INCORRECT
            correct_cell.border = BORDER_CORRECT_ANSWER

    return score, total


# =============================================================================
# Correction d'un fichier complet
# =============================================================================

def correct_test_file(test_file: str, vocab_file: str, sheets: list[str],
                      output_file: str) -> None:
    """
    Corrige le fichier de test et sauvegarde le résultat avec une feuille Correction.
    """
    if not os.path.isfile(test_file):
        print(f"  [ERREUR] Fichier test introuvable : {test_file}")
        return

    answer_key = load_answer_key(vocab_file, sheets)
    wb = openpyxl.load_workbook(test_file)

    scores:  dict[str, int] = {}
    totals:  dict[str, int] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == "Correction":
            continue
        ws = wb[sheet_name]
        score, total = _correct_sheet(ws, answer_key)
        scores[sheet_name] = score
        totals[sheet_name] = total

    # ---- Feuille de bilan ----
    if "Correction" in wb.sheetnames:
        del wb["Correction"]
    bilan = wb.create_sheet("Correction")

    row_idx = 1
    for sheet_name, score in scores.items():
        total = totals[sheet_name]
        if total == 0:
            continue
        pct = f"{(score / total) * 100:.1f}%"
        bilan.cell(row=row_idx, column=1, value=f"Score {sheet_name}")
        bilan.cell(row=row_idx, column=2, value=f"{score} / {total}")
        bilan.cell(row=row_idx, column=3, value=pct)
        row_idx += 2

    total_score = sum(scores.values())
    total_qs    = sum(totals.values())
    if total_qs > 0:
        bilan.cell(row=row_idx, column=1, value="Score total")
        bilan.cell(row=row_idx, column=2, value=f"{total_score} / {total_qs}")
        bilan.cell(row=row_idx, column=3, value=f"{(total_score / total_qs) * 100:.1f}%")

    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    wb.save(output_file)
    print(f"  OK  {os.path.basename(output_file)}  "
          f"(score: {total_score}/{total_qs})")


# =============================================================================
# Point d'entrée
# =============================================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Corrige des tests de vocabulaire japonais."
    )
    parser.add_argument(
        "--profile", "-p",
        choices=list(PROFILES.keys()),
        default=None,
        help="Profil à utiliser.",
    )
    parser.add_argument(
        "--test", "-t",
        nargs="+",
        default=None,
        metavar="SHEET",
        help="Feuilles spécifiques à corriger (ex: N4-14). Par défaut: toutes.",
    )
    return parser.parse_args()


def run_corrections(profile_name: str, profile: dict,
                    sheets_filter: list[str] | None) -> None:
    vocab_file = profile["vocab_file"]
    output_dir = profile["output_dir"]
    level      = profile["level"]

    if not os.path.isfile(vocab_file):
        print(f"[ERREUR] Fichier source introuvable : {vocab_file}")
        return

    # Lister les fichiers _voc_test.xlsx disponibles
    if not os.path.isdir(output_dir):
        print(f"[ERREUR] Dossier de sortie introuvable : {output_dir}")
        return

    test_files = sorted(
        f for f in os.listdir(output_dir) if f.endswith("_voc_test.xlsx")
    )

    if sheets_filter:
        # Ne garder que les fichiers correspondant aux feuilles demandées
        test_files = [
            f for f in test_files
            if any(f.startswith(f"{s}_") or f"{s}_" in f for s in sheets_filter)
        ]

    if not test_files:
        print(f"[AVERTISSEMENT] Aucun fichier *_voc_test.xlsx trouvé dans {output_dir}")
        return

    print(f"\n=== Correction profil : {profile_name} ({len(test_files)} fichier(s)) ===")

    for filename in test_files:
        # Extraire le numéro de feuille depuis le nom de fichier (ex: N4-14_voc_test.xlsx)
        sheet_id = filename.replace("_voc_test.xlsx", "")  # → N4-14
        sheets   = [sheet_id]

        test_path   = os.path.join(output_dir, filename)
        output_path = os.path.join(output_dir, f"{sheet_id}_corrige.xlsx")

        print(f"  Correction de {filename}...", end=" ", flush=True)
        try:
            correct_test_file(test_path, vocab_file, sheets, output_path)
        except Exception as e:
            print(f"\n  [ERREUR] {filename} : {e}")


def main():
    args = parse_args()

    profiles_to_run = (
        {args.profile: PROFILES[args.profile]}
        if args.profile
        else PROFILES
    )

    for profile_name, profile in profiles_to_run.items():
        run_corrections(profile_name, profile, args.test)

    print("\nTerminé.")


if __name__ == "__main__":
    main()
